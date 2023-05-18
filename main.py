# Made by Skies
# Debugged by Skies
# 2023-05-17
import openpyxl
authenticated_user = "skies"
authenticated_password = "nightcoreskies"

def load_database():
  database = []

  # Load the data from the Excel file
  try:
    workbook = openpyxl.load_workbook("flights.xlsx")
    sheet = workbook.active

    # Assuming the data starts from the second row (excluding headers)
    for row in sheet.iter_rows(min_row=2, values_only=True):
      flight = {
        "flight_number": row[0],
        "airline": row[1],
        "origin": row[2],
        "destination": row[3],
        "departure_time": row[4],
        "arrival_time": row[5]
      }
      database.append(flight)

    # Print the loaded database as a table
    if database:
      print("Loaded database:")
      print("{:<15} {:<15} {:<15} {:<15} {:<15} {:<15}".format(
        "Flight Number", "Airline", "Origin", "Destination", "Departure Time",
        "Arrival Time"))
      for flight in database:
        print("{:<15} {:<15} {:<15} {:<15} {:<15} {:<15}".format(
          flight["flight_number"], flight["airline"], flight["origin"],
          flight["destination"], flight["departure_time"],
          flight["arrival_time"]))
    else:
      print("No flights found in the database.")

  except FileNotFoundError:
    print("Database file not found.")

  return database


# Function to save the flight database to the Excel file
def save_database(database):
  wb = openpyxl.Workbook()
  sheet = wb.active
  headers = [
    "Flight Number", "Airline", "Origin", "Destination", "Departure Time",
    "Arrival Time"
  ]
  sheet.append(headers)
  for flight in database:
    row = [
      flight["flight_number"], flight["airline"], flight["origin"],
      flight["destination"], flight["departure_time"], flight["arrival_time"]
    ]
    sheet.append(row)
  wb.save("flights.xlsx")

# def add_flight(database):
#   while True:
#     flight_number = input("Enter the flight number: ").upper()
#
#     # Check if the flight number already exists in the database
#     for flight in database:
#       if flight["flight_number"] == flight_number:
#         print("A flight with the same flight number already exists:")
#         print_flight_details(flight)
#         choice = input("Do you want to update this flight? (Y/N): ").upper()
#         if choice == "Y":
#           update_flight(database)
#           return
#         elif choice == "N":
#           break
#     else:
#       # Continue with adding a new flight
#       break
#
#   while True:
#     airline = input("Enter the airline: ").upper()
#     origin = input("Enter the origin: ").upper()
#     destination = input("Enter the destination: ").upper()
#     departure_time = input("Enter the departure time: ").upper()
#     # Validate the departure time
#     if not is_valid_time(departure_time):
#       print("Invalid departure time. Please provide a valid 24-hour clock time.")
#       continue
#     arrival_time = input("Enter the arrival time: ").upper()
#     # Validate the arrival time
#     if not is_valid_time(arrival_time):
#       print("Invalid arrival time. Please provide a valid 24-hour clock time.")
#       continue
#
#     new_flight = {
#       "flight_number": flight_number,
#       "airline": airline,
#       "origin": origin,
#       "destination": destination,
#       "departure_time": departure_time,
#       "arrival_time": arrival_time
#     }
#
#     # Add the new flight to the database
#     database.append(new_flight)
#
#     # Save the updated database to the Excel file
#     save_database(database)
#
#     print("New flight added successfully!")
#     break

def add_flight(database):
  def is_valid_flight_number(flight_number):
    if len(flight_number) < 6:
      return False
    if not flight_number[:3].isalpha():
      return False
    if not flight_number[3:].isdigit():
      return False
    return True

  def is_valid_location(location):
    if len(location) != 4:
      return False
    if not location.isalpha():
      return False
    return True

  def is_valid_airline(airline):
    if len(airline) != 3:
      return False
    if not airline.isalpha():
      return False
    return True

  flight_number = input("Enter the flight number: ").upper()
  while not is_valid_flight_number(flight_number):
    print("Invalid flight number. Please provide a valid flight number. Ex.CPA520")
    flight_number = input("Enter the flight number: ").upper()

  # Check if the flight number already exists in the database
  for flight in database:
    if flight["flight_number"] == flight_number:
      print("A flight with the same flight number already exists:")
      print_flight_details(flight)
      choice = input("Do you want to update this flight? (Y/N): ").upper()
      if choice == "Y":
        update_flight(database)
        return
      elif choice == "N":
        origin = input("Enter the origin: ").upper()
        while not is_valid_location(origin):
          print("Invalid origin. Please provide a airport ICAO.")
          origin = input("Enter the origin: ").upper()

        destination = input("Enter the destination: ").upper()
        while not is_valid_location(destination):
          print("Invalid destination. Please provide a valid airport ICAO.")
          destination = input("Enter the destination: ").upper()

        departure_time = input("Enter the departure time: ").upper()
        while not is_valid_time(departure_time):
          print("Invalid departure time. Please provide a valid 24-hour clock time. Ex.0300")
          departure_time = input("Enter the departure time: ").upper()

        arrival_time = input("Enter the arrival time: ").upper()
        while not is_valid_time(arrival_time):
          print("Invalid arrival time. Please provide a valid 24-hour clock time. Ex.0300")
          arrival_time = input("Enter the arrival time: ").upper()

        new_flight = {
          "flight_number": flight_number,
          "airline": flight["airline"],
          "origin": origin,
          "destination": destination,
          "departure_time": departure_time,
          "arrival_time": arrival_time
        }

        database.append(new_flight)
        save_database(database)

        print("New flight with the same flight number added successfully!")
        load_database()
        return

  airline = input("Enter the airline: ").upper()
  while not is_valid_airline(airline):
    print("Invalid airline name. Please provide a valid airline ICAO.")
    airline = input("Enter the airline: ").upper()

  origin = input("Enter the origin: ").upper()
  while not is_valid_location(origin):
    print("Invalid origin. Please provide a valid airport ICAO.")
    origin = input("Enter the origin: ").upper()

  destination = input("Enter the destination: ").upper()
  while not is_valid_location(destination):
    print("Invalid destination. Please provide a valid airport ICAO")
    destination = input("Enter the destination: ").upper()

  departure_time = input("Enter the departure time: ").upper()
  while not is_valid_time(departure_time):
    print("Invalid departure time. Please provide a valid 24-hour clock time. Ex.0300")
    departure_time = input("Enter the departure time: ").upper()

  arrival_time = input("Enter the arrival time: ").upper()
  while not is_valid_time(arrival_time):
    print("Invalid arrival time. Please provide a valid 24-hour clock time. Ex.0300")
    arrival_time = input("Enter the arrival time: ").upper()

  new_flight = {
    "flight_number": flight_number,
    "airline": airline,
    "origin": origin,
    "destination": destination,
    "departure_time": departure_time,
    "arrival_time": arrival_time
  }

  database.append(new_flight)
  save_database(database)

  print("New flight added successfully!")
  load_database(database)

def is_valid_time(time_str):
  if len(time_str) != 4:  # Check if the length is not 4 characters
    return False

  try:
    hours = int(time_str[:2])
    minutes = int(time_str[2:])
    if hours < 0 or hours >= 24 or minutes < 0 or minutes >= 60:
      return False
  except ValueError:
    return False

  return True

def save_database(database):
  workbook = openpyxl.Workbook()
  sheet = workbook.active

  # Write headers
  headers = [
    "Flight Number", "Airline", "Origin", "Destination", "Departure Time",
    "Arrival Time"
  ]
  sheet.append(headers)

  # Write data
  for flight in database:
    row = [
      flight["flight_number"], flight["airline"], flight["origin"],
      flight["destination"], flight["departure_time"], flight["arrival_time"]
    ]
    sheet.append(row)

  # Save the workbook to the Excel file
  workbook.save("flights.xlsx")

  print("Database saved successfully!")


# Function to remove a flight from the database
def remove_flight(database):
  flight_number = input("Enter the flight number of the flight to remove: ")

  # Find all flights with the given flight number
  matching_flights = [flight for flight in database if flight['flight_number'] == flight_number]

  if not matching_flights:
    print("No flights found with the given flight number.")
    return

  if len(matching_flights) == 1:
    # Only one flight found, remove it
    database.remove(matching_flights[0])
    print("Flight successfully removed.")
    load_database(database)
    save_database(database)
    return

  # Display the flights with the same flight number
  print("Multiple flights found with the same flight number:")
  for i, flight in enumerate(matching_flights):
    print(f"{i + 1}. {flight['flight_number']}, {flight['airline']}, {flight['origin']}, {flight['destination']}, "
          f"Departure: {flight['departure_time']}, Arrival: {flight['arrival_time']}")

  # Prompt the user to choose which entry to remove
  while True:
    choice = input("Enter the number of the flight to remove: ")
    try:
      choice = int(choice)
      if 1 <= choice <= len(matching_flights):
        break
      else:
        print("Invalid choice. Please enter a valid number.")
    except ValueError:
      print("Invalid choice. Please enter a valid number.")

  # Remove the chosen flight
  chosen_flight = matching_flights[choice - 1]
  database.remove(chosen_flight)
  print("Flight successfully removed.")

# Function to update flight information
def update_flight(database):
    flight_number = input("Enter the flight number to update: ").upper()

    # Find the flights with the given flight number
    matching_flights = [flight for flight in database if flight["flight_number"] == flight_number]

    if not matching_flights:
        print("Flight not found in the database.")
        return

    # Display the flights with the same flight number
    print("Multiple flights found with the same flight number:")
    for i, flight in enumerate(matching_flights):
        print(f"{i + 1}. {flight['flight_number']}, {flight['airline']}, {flight['origin']}, {flight['destination']}, "
              f"Departure: {flight['departure_time']}, Arrival: {flight['arrival_time']}")

    # Prompt the user to choose which entry to modify
    while True:
        choice = input("Enter the number of the flight to update: ")
        try:
            choice = int(choice)
            if 1 <= choice <= len(matching_flights):
                break
            else:
                print("Invalid choice. Please enter a valid number.")
        except ValueError:
            print("Invalid choice. Please enter a valid number.")

    # Select the chosen flight
    chosen_flight = matching_flights[choice - 1]

    field = input(
        "Enter the field to update (airline, origin, destination, departure time, arrival time): "
    ).lower()
    new_value = input("Enter the new value: ").upper()

    # Update the specified field with the new value
    if field == "airline":
        chosen_flight["airline"] = new_value
    elif field == "origin":
        chosen_flight["origin"] = new_value
    elif field == "destination":
        chosen_flight["destination"] = new_value
    elif field == "departure time":
        chosen_flight["departure_time"] = new_value
    elif field == "arrival time":
        chosen_flight["arrival_time"] = new_value
    else:
        print("Invalid field. Please try again.")
        return

    # Save the updated database to the Excel file
    save_database(database)

    print("Flight information updated successfully!")
    load_database()

# Function to sort flights based on user choice
def sort_flights(database):
    sort_option = input("Enter the sort option (flight_number/airline/origin/destination): ").lower()

    # Check if the sort option is valid
    if sort_option not in ["flight_number", "airline", "origin", "destination"]:
      print("Invalid sort option.")
      return

    # Sort the database based on the chosen sort option
    sorted_database = sorted(database, key=lambda flight: flight[sort_option])

    # Print the sorted flights as a table
    print("{:<15} {:<15} {:<15} {:<15} {:<15} {:<15}".format(
      "Flight Number", "Airline", "Origin", "Destination", "Departure Time",
      "Arrival Time"))
    for flight in sorted_database:
      print("{:<15} {:<15} {:<15} {:<15} {:<15} {:<15}".format(
        flight["flight_number"], flight["airline"], flight["origin"],
        flight["destination"], flight["departure_time"], flight["arrival_time"]))


# Function to search for a flight based on a keyword
def search_flight(database):
  keyword = input("Enter the keyword to search: ")
  matching_flights = []
  for flight in database:
    if (keyword.lower() in flight["flight_number"].lower()
        or keyword.lower() in flight["airline"].lower()
        or keyword.lower() in flight["origin"].lower()
        or keyword.lower() in flight["destination"].lower()
        or keyword.lower() in flight["departure_time"].lower()
        or keyword.lower() in flight["arrival_time"].lower()):
      matching_flights.append(flight)
  if matching_flights:
    print("Matching flights:")
    for flight in matching_flights:
      print(flight)
  else:
    print("No flights found matching the keyword.")


# Function to print flight details
def print_flight_details(flight):
  print("Flight Number:", flight["flight_number"])
  print("Airline:", flight["airline"])
  print("Origin:", flight["origin"])
  print("Destination:", flight["destination"])
  print("Departure Time:", flight["departure_time"])
  print("Arrival Time:", flight["arrival_time"])


# Main menu loop
# def sort_flights(database):
#   print("Select the criteria to sort by:")
#   print("1. Flight Number")
#   print("2. Airline")
#   print("3. Origin")
#   print("4. Destination")
#   print("5. Departure Time")
#   print("6. Arrival Time")
#
#   choice = input("Enter your choice: ")
#
#   if choice == "1":
#     sorted_flights = sorted(database, key=lambda x: x["flight_number"])
#   elif choice == "2":
#     sorted_flights = sorted(database, key=lambda x: x["airline"])
#   elif choice == "3":
#     sorted_flights = sorted(database, key=lambda x: x["origin"])
#   elif choice == "4":
#     sorted_flights = sorted(database, key=lambda x: x["destination"])
#   elif choice == "5":
#     sorted_flights = sorted(database, key=lambda x: x["departure_time"])
#   elif choice == "6":
#     sorted_flights = sorted(database, key=lambda x: x["arrival_time"])
#   else:
#     print("Invalid choice.")
#     return
#
#   print("Sorted flights:")
#   for flight in sorted_flights:
#     print(flight)

# Function to search for a flight based on a keyword
def search_flight(database):
  keyword = input("Enter the keyword to search: ")
  matching_flights = []

  for flight in database:
    if (keyword.lower() in flight["flight_number"].lower()
        or keyword.lower() in flight["airline"].lower()
        or keyword.lower() in flight["origin"].lower()
        or keyword.lower() in flight["destination"].lower()
        or keyword.lower() in flight["departure_time"].lower()
        or keyword.lower() in flight["arrival_time"].lower()):
      matching_flights.append(flight)

  if matching_flights:
    print("Matching flights:")
    for flight in matching_flights:
      print(flight)
  else:
    print("No flights found matching the keyword.")

def delete_all_entries(database):
    confirm = input("Are you sure you want to delete all entries? (Y/N): ").upper()
    if confirm == "Y":
      database.clear()
      save_database(database)
      print("All entries deleted successfully!")
    elif confirm == "N":
      print("Deletion canceled.")

updates = []

# Function to load updates from file
def load_updates():
    with open("updates.txt", "r") as file:
        updates = file.readlines()
    return updates

# Function to add an update
def add_update():
    # Check if the user is authenticated
    username = input("Enter your username: ")
    password = input("Enter your password: ")
    if username != authenticated_user or password != authenticated_password:
        print("Access denied.")
        return

    update = input("Enter the update: ")

    # Append the update to the updates list
    updates.append(update + "\n")

    # Save the updates to the file
    with open("updates.txt", "a") as file:
        file.write(update + "\n")

    print("Update added successfully.")

# Function to view updates
def view_updates():
    # Load updates from the file
    updates = load_updates()

    print("Recent Updates:")
    for update in updates:
        print("- " + update.strip())

    # Check if the user is authenticated
    if input("Login to add update? (Y/N): ").upper() == "Y":
        # Call the add_update function
        add_update()

# Main menu loop
def main_menu(database):
  while True:
    print("\nFlight Database Menu:")
    print("1. Add a flight")
    print("2. Remove a flight")
    print("3. Change flight information")
    print("4. Sort flights")
    print("5. Search for a flight")
    print("6. Save database")
    print("7. Load database")
    print("8. Clear all flight entries")
    print("9. View updates")
    print("10. Exit")

    choice = input("Enter your choice: ")

    if choice == "1":
      add_flight(database)
    elif choice == "2":
      remove_flight(database)
    elif choice == "3":
      update_flight(database)
    elif choice == "4":
      sort_flights(database)
    elif choice == "5":
      search_flight(database)
    elif choice == "6":
      save_database(database)
      print("Database saved successfully!")
      load_database()
    elif choice == "7":
      database = load_database()
      print("Database loaded successfully!")
    elif choice == "8":
      delete_all_entries(database)
    elif choice == "9":
      view_updates()
    elif choice == "10":
      print("Exiting...")
      break
    else:
      print("Invalid choice. Please try again.")


# Load the initial flight database
flight_database = load_database()

# Run the main menu
main_menu(flight_database)